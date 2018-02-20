"""Excel module"""
class VaultExcel :

    def __init__(self) :
        pass



    def __before_create_project(workdir):
        # Debug output
        print("before create project")
        
        import pathlib
        pathlib.Path(workdir).mkdir(parents=True, exist_ok=True)



    def create_project(workdir):
        VaultExcel.__before_create_project(workdir)
    
        from openpyxl import Workbook
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        ws.title = "V1.Changelog"
        ws.sheet_properties.tabColor = "007700"
        
        # W sheets
        wb.create_sheet("W1.Information")
        ws2 = wb["W1.Information"]
        ws2.sheet_properties.tabColor = "118800"
        
        # X sheets
        wb.create_sheet("X1.Tables")
        ws3 = wb["X1.Tables"]
        ws3.sheet_properties.tabColor = "229900"
        
        # Y sheets
        wb.create_sheet("Y1.Fields")
        ws4 = wb["Y1.Fields"]
        ws4.sheet_properties.tabColor = "33AA00"
        
        # Z sheets
        wb.create_sheet("Z1.Examples")
        ws5 = wb["Z1.Examples"]
        ws5.sheet_properties.tabColor = "44BB00"

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()

        # Save the file
        import os
        wb.save(os.path.join(workdir, "01_document.xlsx"))

        # Debug output
        VaultExcel.__after_create_project(workdir)
        print("ok")


    
    def __after_create_project(workdir):
        from openpyxl import Workbook
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()

        # Save the file
        import os
        wb.save(os.path.join(workdir, "01_example.xlsx"))

        # Debug output
        print("after create project")



if __name__ == "__main__":
    VaultExcel.create_project(".")
